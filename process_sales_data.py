""" 
Description: 
  Divides sales data CSV file into individual order data Excel files.

Usage:
  python process_sales_data.py sales_csv_path

Parameters:
  sales_csv_path = Full path of the sales data CSV file
"""
import pandas as pd
from sys import argv
import re
import os
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

def main():
    sales_csv_path = get_sales_csv_path()
    orders_dir_path = create_orders_dir(sales_csv_path)
    process_sales_data(sales_csv_path, orders_dir_path)

def get_sales_csv_path():    
    """Gets the path of sales data CSV file from the command line

    Returns:
        str: Path of sales data CSV file
    """
    #Check whether command line parameter provided
    try:
        file_path = argv[1]
    except IndexError:
        print("Error: Please provide the file path")
        quit()
    #Check whether provide parameter is valid path of file
    try:
        open(file_path, 'r')
    except:
        print("Error: File not found")
        quit()
    #Return path of sales data CSV file
    return file_path

def create_orders_dir(sales_csv_path):
    """Creates the directory to hold the individual order Excel sheets

    Args:
        sales_csv_path (str): Path of sales data CSV file

    Returns:
        str: Path of orders directory
    """
    #Get directory in which sales data CSV file resides
    salesParentDirectory = os.path.abspath(os.path.join(sales_csv_path, os.pardir))
    #Determine the path of the directory to hold the order data files
    match = re.search(r".*^\S+", str(datetime.now()))
    #Create the orders directory if it does not already exist
    orderDir = f"{salesParentDirectory}\\Orders_{match.group()}"
    try:
        os.mkdir(orderDir)
    except OSError:
        pass
    #Return path of orders directory
    return orderDir

def process_sales_data(sales_csv_path, orders_dir_path):
    """Splits the sales data into individual orders and save to Excel sheets

    Args: 
        sales_csv_path (str): Path of sales data CSV file
        orders_dir_path (str): Path of orders directory
    """
    #Import the sales data from the CSV file into a DataFrame
    salesDataFrame = pd.read_csv(sales_csv_path)
    
    #Insert a new "TOTAL PRICE" column into the DataFrame
    salesDataFrame.insert(7, 'TOTAL PRICE', salesDataFrame['ITEM QUANTITY'] * salesDataFrame['ITEM PRICE'])

    #Remove columns from the DataFrame that are not needed
    salesDataFrame = salesDataFrame.drop(['ADDRESS', 'CITY', 'STATE', 'POSTAL CODE', 'COUNTRY'], axis=1)

    #Dropping duplicate ORDER ID's
    salesDataFrame = salesDataFrame.drop_duplicates(subset=['ORDER ID'])

    #Remove the 'ORDER ID' column
    salesDataFrame = salesDataFrame.drop(['ORDER ID'], axis=1)

    #Sort the items by item number
    salesDataFrame = salesDataFrame.sort_values('ITEM NUMBER')

    #Calculate the TOTAL PRICE
    totalPrice = salesDataFrame['TOTAL PRICE'].sum()
    
    #Determine the file name and full path of the Excel sheet
    excelFilePath = f'{orders_dir_path}\\sales_csv.xlsx'

    #Export the data to an Excel sheet
    try:
        salesDataFrame.to_excel(excelFilePath, index=False, sheet_name='Sales Info')
    except PermissionError:
        print("ERROR: FILE WITH SAME NAME IS OPEN")
    
    #Load in the excel file and sheet using openpyxl
    workbook = load_workbook(excelFilePath)
    worksheet = workbook.active

    #Creating the grand total row using empty strings and the total price then appending it
    grandTotalRow = ['','','','','','GRAND TOTAL', totalPrice,'','','']
    worksheet.append(grandTotalRow)

    #Format each colunm then saving changes
    for column_cells in worksheet.columns:
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = max(len(str(cell.value) or "") for cell in column_cells)
    
    #Define format for the money columns
    for col in worksheet.iter_cols(min_row=1, min_col=6, max_col=7):
        for cell in col:
            cell.number_format = '$#,##0.00'  

    #Format the Grand Total row
    worksheet[f'G{worksheet.max_row}'].number_format = '$#,##0.00'

    workbook.save(excelFilePath)

    return

if __name__ == '__main__':
    main()